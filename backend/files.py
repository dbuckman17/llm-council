"""Per-conversation file upload, storage, and text extraction."""

import base64
import json
import os
import uuid
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

from .config import DATABASE_URL

FILES_DIR = "data/files"
USE_POSTGRES = bool(DATABASE_URL)

# Max image size for base64 pass-through (2 MB)
MAX_IMAGE_BYTES = 2 * 1024 * 1024

IMAGE_CONTENT_TYPES = {
    "image/png", "image/jpeg", "image/jpg", "image/gif", "image/webp",
}

TEXT_CONTENT_TYPES = {
    "text/plain", "text/csv", "text/markdown", "text/html",
    "application/json", "application/xml",
}


@dataclass
class ConversationFile:
    id: str
    conversation_id: str
    filename: str
    content_type: str
    size_bytes: int
    extracted_text: Optional[str]
    is_image: bool
    storage_path: str
    created_at: str


def _files_dir(conversation_id: str) -> Path:
    p = Path(FILES_DIR) / conversation_id
    p.mkdir(parents=True, exist_ok=True)
    return p


def _manifest_path(conversation_id: str) -> Path:
    return _files_dir(conversation_id) / "_manifest.json"


def _load_manifest(conversation_id: str) -> List[dict]:
    mp = _manifest_path(conversation_id)
    if mp.exists():
        with open(mp) as f:
            return json.load(f)
    return []


def _save_manifest(conversation_id: str, entries: List[dict]):
    mp = _manifest_path(conversation_id)
    mp.parent.mkdir(parents=True, exist_ok=True)
    with open(mp, "w") as f:
        json.dump(entries, f, indent=2)


def _extract_text(file_path: str, content_type: str, filename: str) -> Optional[str]:
    """Extract text content from a file based on its type."""
    lower = filename.lower()

    # Plain text / code / CSV / JSON / Markdown
    if content_type in TEXT_CONTENT_TYPES or lower.endswith(
        (".py", ".js", ".ts", ".jsx", ".tsx", ".go", ".rs", ".java",
         ".c", ".cpp", ".h", ".rb", ".swift", ".kt", ".sh", ".yaml",
         ".yml", ".toml", ".ini", ".cfg", ".sql", ".r", ".m", ".cs")
    ):
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        except Exception:
            return None

    # PDF
    if lower.endswith(".pdf") or content_type == "application/pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            pages = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages) if pages else None
        except Exception as e:
            print(f"PDF extraction failed for {filename}: {e}")
            return None

    # DOCX
    if lower.endswith(".docx") or content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        try:
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs if p.text)
        except Exception as e:
            print(f"DOCX extraction failed for {filename}: {e}")
            return None

    # XLSX
    if lower.endswith(".xlsx") or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(rows) if rows else None
        except Exception as e:
            print(f"XLSX extraction failed for {filename}: {e}")
            return None

    return None


async def save_file(
    conversation_id: str,
    filename: str,
    content_type: str,
    file_data: bytes,
) -> ConversationFile:
    """Save an uploaded file and extract text if possible."""
    file_id = str(uuid.uuid4())
    safe_name = filename.replace("/", "_").replace("\\", "_")
    storage_path = str(_files_dir(conversation_id) / f"{file_id}_{safe_name}")

    # Write file to disk
    with open(storage_path, "wb") as f:
        f.write(file_data)

    is_image = content_type in IMAGE_CONTENT_TYPES
    extracted_text = None if is_image else _extract_text(storage_path, content_type, filename)

    cf = ConversationFile(
        id=file_id,
        conversation_id=conversation_id,
        filename=filename,
        content_type=content_type,
        size_bytes=len(file_data),
        extracted_text=extracted_text,
        is_image=is_image,
        storage_path=storage_path,
        created_at=datetime.utcnow().isoformat(),
    )

    if USE_POSTGRES:
        from .db import get_pool
        pool = await get_pool()
        await pool.execute(
            "INSERT INTO conversation_files (id, conversation_id, filename, content_type, "
            "size_bytes, extracted_text, is_image, storage_path, created_at) "
            "VALUES ($1, $2, $3, $4, $5, $6, $7, $8, NOW())",
            cf.id, cf.conversation_id, cf.filename, cf.content_type,
            cf.size_bytes, cf.extracted_text, cf.is_image, cf.storage_path,
        )
    else:
        entries = _load_manifest(conversation_id)
        entries.append(asdict(cf))
        _save_manifest(conversation_id, entries)

    return cf


async def get_conversation_files(conversation_id: str) -> List[dict]:
    """List files for a conversation."""
    if USE_POSTGRES:
        from .db import get_pool
        pool = await get_pool()
        rows = await pool.fetch(
            "SELECT id, filename, content_type, size_bytes, is_image, created_at "
            "FROM conversation_files WHERE conversation_id = $1 ORDER BY created_at",
            conversation_id,
        )
        return [
            {
                "id": r["id"],
                "filename": r["filename"],
                "content_type": r["content_type"],
                "size_bytes": r["size_bytes"],
                "is_image": r["is_image"],
                "created_at": r["created_at"].isoformat() if hasattr(r["created_at"], "isoformat") else r["created_at"],
            }
            for r in rows
        ]

    entries = _load_manifest(conversation_id)
    return [
        {
            "id": e["id"],
            "filename": e["filename"],
            "content_type": e["content_type"],
            "size_bytes": e["size_bytes"],
            "is_image": e["is_image"],
            "created_at": e["created_at"],
        }
        for e in entries
    ]


async def delete_file(conversation_id: str, file_id: str) -> bool:
    """Delete a file by ID."""
    if USE_POSTGRES:
        from .db import get_pool
        pool = await get_pool()
        row = await pool.fetchrow(
            "SELECT storage_path FROM conversation_files WHERE id = $1 AND conversation_id = $2",
            file_id, conversation_id,
        )
        if row is None:
            return False
        if os.path.exists(row["storage_path"]):
            os.remove(row["storage_path"])
        await pool.execute(
            "DELETE FROM conversation_files WHERE id = $1",
            file_id,
        )
        return True

    entries = _load_manifest(conversation_id)
    new_entries = []
    found = False
    for e in entries:
        if e["id"] == file_id:
            found = True
            if os.path.exists(e["storage_path"]):
                os.remove(e["storage_path"])
        else:
            new_entries.append(e)
    if found:
        _save_manifest(conversation_id, new_entries)
    return found


async def get_file_path(conversation_id: str, file_id: str) -> Optional[str]:
    """Get storage path for a file (for download)."""
    if USE_POSTGRES:
        from .db import get_pool
        pool = await get_pool()
        row = await pool.fetchrow(
            "SELECT storage_path, filename FROM conversation_files WHERE id = $1 AND conversation_id = $2",
            file_id, conversation_id,
        )
        if row and os.path.exists(row["storage_path"]):
            return row["storage_path"]
        return None

    entries = _load_manifest(conversation_id)
    for e in entries:
        if e["id"] == file_id and os.path.exists(e["storage_path"]):
            return e["storage_path"]
    return None


async def get_file_content_for_context(conversation_id: str) -> Tuple[str, List[dict]]:
    """
    Build context from all conversation files.

    Returns:
        Tuple of (text_context, image_attachments)
        - text_context: concatenated extracted text from all text files
        - image_attachments: list of {mime_type, base64_data} for vision pass-through
    """
    text_parts = []
    image_attachments = []

    if USE_POSTGRES:
        from .db import get_pool
        pool = await get_pool()
        rows = await pool.fetch(
            "SELECT filename, extracted_text, is_image, storage_path, content_type, size_bytes "
            "FROM conversation_files WHERE conversation_id = $1 ORDER BY created_at",
            conversation_id,
        )
        for r in rows:
            if r["is_image"]:
                if r["size_bytes"] <= MAX_IMAGE_BYTES and os.path.exists(r["storage_path"]):
                    with open(r["storage_path"], "rb") as f:
                        b64 = base64.b64encode(f.read()).decode("utf-8")
                    image_attachments.append({
                        "mime_type": r["content_type"],
                        "base64_data": b64,
                    })
            elif r["extracted_text"]:
                text_parts.append(f"--- {r['filename']} ---\n{r['extracted_text']}")
    else:
        entries = _load_manifest(conversation_id)
        for e in entries:
            if e.get("is_image"):
                if e["size_bytes"] <= MAX_IMAGE_BYTES and os.path.exists(e["storage_path"]):
                    with open(e["storage_path"], "rb") as f:
                        b64 = base64.b64encode(f.read()).decode("utf-8")
                    image_attachments.append({
                        "mime_type": e["content_type"],
                        "base64_data": b64,
                    })
            elif e.get("extracted_text"):
                text_parts.append(f"--- {e['filename']} ---\n{e['extracted_text']}")

    text_context = "\n\n".join(text_parts) if text_parts else ""
    return text_context, image_attachments
